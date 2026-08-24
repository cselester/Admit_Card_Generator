"""
Fast admit card generator — PDF overlay approach.

Instead of the old flow (copy Google Doc -> find/replace -> export PDF -> upload),
this stamps student data directly onto a pre-rendered background PDF page,
entirely in-process, with no network calls per student.

Usage:
    python3 generate_admit_cards.py --sheet CBT --limit 200 --workers 8
"""

import argparse
import io
import time
from pathlib import Path
from concurrent.futures import ProcessPoolExecutor, as_completed

import openpyxl
from pypdf import PdfReader, PdfWriter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase.pdfmetrics import stringWidth

TEMPLATE_PDF = Path(__file__).parent / "template_bg.pdf"
OUTPUT_DIR = Path(__file__).parent / "output_cards"
FONT = "Helvetica"
FONT_SIZE = 11
TEXT_X = 166          # left edge of the value column text (matches template)
CELL_X0, CELL_X1 = 160, 560   # whiteout rectangle spans the value cell (kept clear of the 575.5pt border)
PAGE_W, PAGE_H = 612, 792     # US Letter, matches template render

# Field map: derived once from the template's placeholder positions.
# top/bottom are in "top-down" PDF coordinates (as reported by pdfplumber);
# converted to reportlab's bottom-up coordinates at draw time.
FIELD_MAP = {
    "Registration Number": {"cell_top": 166.5, "cell_bottom": 192.5, "text_bottom": 184.64},
    "Student Name":         {"cell_top": 192.5, "cell_bottom": 217.5, "text_bottom": 210.14},
    "Registration Status":  {"cell_top": 217.5, "cell_bottom": 243.5, "text_bottom": 235.64},
    "Center":                {"cell_top": 243.5, "cell_bottom": 268.5, "text_bottom": 261.14},
    "CBT Center":            {"cell_top": 268.5, "cell_bottom": 294.5, "text_bottom": 286.64},
    "CBT Center Address":    {"cell_top": 294.5, "cell_bottom": 331.5, "text_bottom": 312.14},
    "CBT Exam Timings":      {"cell_top": 331.5, "cell_bottom": 357.5, "text_bottom": 349.64},
}

# Table grid geometry, read once off the template — redrawn fresh on every
# card instead of relying on the original lines surviving the PDF merge
# (some viewers render the merged/underlying strokes inconsistently).
GRID_LEFT_X = 36.5
GRID_MID_X = 153.5      # divider between label column and value column
GRID_RIGHT_X = 574.5
GRID_ROW_YS = [166.5, 192.5, 217.5, 243.5, 268.5, 294.5, 331.5, 357.5]


def draw_grid(c):
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(0.75)
    for y in GRID_ROW_YS:
        yy = PAGE_H - y
        c.line(GRID_LEFT_X, yy, GRID_RIGHT_X, yy)
    c.line(GRID_LEFT_X, PAGE_H - GRID_ROW_YS[0], GRID_LEFT_X, PAGE_H - GRID_ROW_YS[-1])
    c.line(GRID_MID_X, PAGE_H - GRID_ROW_YS[0], GRID_MID_X, PAGE_H - GRID_ROW_YS[-1])
    c.line(GRID_RIGHT_X, PAGE_H - GRID_ROW_YS[0], GRID_RIGHT_X, PAGE_H - GRID_ROW_YS[-1])


def wrap_text(text, max_width, font=FONT, size=FONT_SIZE):
    """Greedy word-wrap so long values (e.g. addresses) don't overrun the cell."""
    words = str(text).split()
    lines, current = [], ""
    for w in words:
        trial = (current + " " + w).strip()
        if stringWidth(trial, font, size) <= max_width:
            current = trial
        else:
            if current:
                lines.append(current)
            current = w
    if current:
        lines.append(current)
    return lines or [""]


def build_overlay(record: dict) -> "io.BytesIO":
    """Draw one student's data as a transparent overlay page."""
    buf = io.BytesIO()
    c = canvas.Canvas(buf, pagesize=(PAGE_W, PAGE_H))
    c.setFont(FONT, FONT_SIZE)

    for field, pos in FIELD_MAP.items():
        raw = record.get(field, "")
        if isinstance(raw, float) and raw.is_integer():
            value = str(int(raw))
        else:
            value = str(raw or "")
        cell_top_y = PAGE_H - pos["cell_top"]
        cell_bottom_y = PAGE_H - pos["cell_bottom"]
        cell_height = cell_top_y - cell_bottom_y
        max_width = CELL_X1 - TEXT_X - 4

        # Whiteout the placeholder text first
        c.setFillColorRGB(1, 1, 1)
        c.rect(CELL_X0, cell_bottom_y, CELL_X1 - CELL_X0, cell_height, fill=1, stroke=0)
        c.setFillColorRGB(0, 0, 0)

        lines = wrap_text(value, max_width)
        if len(lines) == 1:
            y = PAGE_H - pos["text_bottom"]
            c.drawString(TEXT_X, y, lines[0])
        else:
            # Multi-line values: center the block vertically in the cell
            line_gap = FONT_SIZE + 2
            block_height = line_gap * len(lines)
            start_y = cell_bottom_y + (cell_height + block_height) / 2 - FONT_SIZE
            for i, line in enumerate(lines):
                c.drawString(TEXT_X, start_y - i * line_gap, line)

    draw_grid(c)

    c.save()
    buf.seek(0)
    return buf


def render_one(record: dict, out_dir: Path) -> str:
    overlay_buf = build_overlay(record)
    overlay_reader = PdfReader(overlay_buf)
    base_reader = PdfReader(str(TEMPLATE_PDF))

    writer = PdfWriter()
    base_page = base_reader.pages[0]
    base_page.merge_page(overlay_reader.pages[0])
    writer.add_page(base_page)

    reg_no = str(record.get("Registration Number", "unknown")).replace(".0", "")
    out_path = out_dir / f"{reg_no}.pdf"
    with open(out_path, "wb") as f:
        writer.write(f)
    return str(out_path)


def load_records(xlsx_path: Path, sheet_name: str, limit: int | None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rec = dict(zip(headers, row))
        records.append(rec)
        if limit and len(records) >= limit:
            break
    return records


def write_manifest(records: list, out_dir: Path, sheet_name: str):
    """Write a manifest (reg no -> student details -> filename) alongside the PDFs.
    This is what the Drive-side Apps Script reads to enrich the Index sheet,
    and it's a plain local backup of the batch even before anything reaches Drive."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Manifest"
    headers = ["Registration Number", "Student Name", "Batch Code", "Registration Status", "Center",
               "CBT Center", "CBT Center Address", "CBT Exam Timings", "File Name", "Source Sheet"]
    ws.append(headers)
    for rec in records:
        reg_no = rec.get("Registration Number", "")
        if isinstance(reg_no, float) and reg_no.is_integer():
            reg_no = int(reg_no)
        ws.append([
            reg_no,
            rec.get("Student Name", ""),
            rec.get("Batch Code", ""),
            rec.get("Registration Status", ""),
            rec.get("Center", ""),
            rec.get("CBT Center", ""),
            rec.get("CBT Center Address", ""),
            rec.get("CBT Exam Timings", ""),
            f"{reg_no}.pdf",
            sheet_name,
        ])
    manifest_path = out_dir / "batch_manifest.xlsx"
    wb.save(manifest_path)
    return manifest_path



def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="Warrior_Admit_card_Maker_Tool____MV.xlsx")
    ap.add_argument("--sheet", default="CBT")
    ap.add_argument("--limit", type=int, default=200, help="0 = all rows")
    ap.add_argument("--workers", type=int, default=8)
    ap.add_argument("--output-dir", default=None,
                     help="Where to write PDFs. Point this at your Google Drive Desktop "
                          "synced folder to get them into Drive automatically.")
    args = ap.parse_args()

    out_dir = Path(args.output_dir) if args.output_dir else OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    records = load_records(Path(args.xlsx), args.sheet, args.limit or None)
    print(f"Loaded {len(records)} records from sheet '{args.sheet}'")

    start = time.time()
    done = 0
    with ProcessPoolExecutor(max_workers=args.workers) as pool:
        futures = [pool.submit(render_one, rec, out_dir) for rec in records]
        for f in as_completed(futures):
            f.result()
            done += 1

    manifest_path = write_manifest(records, out_dir, args.sheet)

    elapsed = time.time() - start
    rate = done / elapsed if elapsed > 0 else 0
    print(f"\nGenerated {done} admit cards in {elapsed:.2f}s  ({rate:.1f} cards/sec)")
    print(f"Manifest written to {manifest_path}")
    if rate > 0:
        for n in (30000, 100000):
            print(f"  Projected time for {n:,} records: {n / rate / 60:.1f} minutes")


if __name__ == "__main__":
    main()
